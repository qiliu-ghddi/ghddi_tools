"""
Display SMILES in excel
Author: Qi Liu

Generate adjustable table 
0. convert SMILES to image
1. convert image to b64
2. inster image to excel
3. adjust every cell to fit the size of the image
"""
import io
import os
import base64
import shutil
import argparse
import numpy as np
import pandas as pd
import openpyxl
import PIL
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XSImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
from rdkit import Chem
from rdkit.Chem import Draw


def mol_to_png(mol, save_dir='.', basename="test"):
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    filename = f"{save_dir}/{basename}.png"
    Draw.MolToFile(mol, filename)
    return filename


def mols_to_pngs(mols, save_dir='.', basename="test"):
    """Helper to write RDKit mols to png files."""

    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    filenames = []
    for i, mol in enumerate(mols):
        filename = f"{save_dir}/{basename}_{i}.png"
        Draw.MolToFile(mol, filename)
        filenames.append(filename)
    return filenames


def cvt_to_base64(file_name):
    with open(file_name, "rb") as image_file:
        data = base64.b64encode(image_file.read())
    return data.decode('utf-8')


def convert_smiles_base64(smiles, save_dir='.temp', delete=False):
    try:
        mol = Chem.MolFromSmiles(smiles)
        image_fn = mol_to_png(mol, save_dir, smiles)
        base64_str = cvt_to_base64(image_fn)
    except Exception as e:
        print(e)
        base64_str = ""
        
    if len(base64_str)==0:
        image_base64 = ""
    else:
        image_base64 = f'<img src="data:image/png;base64,{base64_str}">'
    
    if delete:
        shutil.rmtree(save_dir)
    
    return image_base64


def convert_base64_to_image(image_name, image_base64, save_dir='.temp', delete=False):
    file = Path(f"{save_dir}/{image_name}")
    if not file.parent.exists():
        file.parent.mkdir()
    
    with open(file, 'wb') as fout:
        bytes_b64 = base64.b64decode(image_base64)
        fout.write(bytes_b64)
    return file


def add_molecular_column_to_frame(df, smiles_col="SMILES", mol_col="ROMol"):
    """add a column of base64 string"""
    df[mol_col] = df[smiles_col][:].apply(lambda x: convert_smiles_base64(x))
    return df


def convert_base64_to_bytes(s):
    b64 = s.split(',')[1]
    bytes_b64 = base64.b64decode(b64)
    bytes_img = io.BytesIO(bytes_b64)
    return bytes_img



def main(args):
    file = Path(args.file)

    smiles_col = "Ligand SMILES"
    mol_col = "ROMol"
    mol_img_col = "ROMolImage"
    
    df = pd.read_csv(str(file), index_col=[0])
    df = add_molecular_column_to_frame(df, smiles_col=smiles_col, mol_col=mol_col)
    # df.to_csv(f"{file.parent}/{file.stem}-base64.csv", index=False)

    # Generate images
    fn_list = []
    for i, tag in df[mol_col].iteritems():
        try:
            soup = BeautifulSoup(tag, 'html.parser')
            bs64 = soup.img['src'].split(',')[1]  # TODO: need to check
            fn = convert_base64_to_image(image_name=f"{file.stem}-{i}.png", image_base64=bs64, save_dir=f"{file.parent}/images")
        except Exception as e:
            print(f"i={i} Error!", e)
            fn = ""
        finally:
            fn_list.append(fn)
            
    # Create and write to excel
    wb = Workbook()
    ws = wb.worksheets[0]
    al = Alignment(horizontal='general', vertical='top')
    
    ws.append(df.columns.tolist())
    for i, row in df.iterrows():
        ws.append(row.tolist())

    num_row, num_col = df.shape
    img_width, img_height = 300, 300

    # resize cells
    # 95.25 value was found here https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/utils/units.html
    # my screen resolution is 1920*1080
    WIDTH_FUDGE = 1080/1920*(1 - 72/95.25)
    HEIGHT_FUDGE = 3/4  # 96 is my screen dpi, 72 is a standart dpi

    # write header
    flag_header = True
    img_col_letter = get_column_letter(num_col+1)
    ws.column_dimensions[img_col_letter].width = img_width*WIDTH_FUDGE
    ws[img_col_letter+str(1)] = mol_img_col
    index_content_start = 2 if flag_header else 1
    index_content_end = num_row+index_content_start
    # write contents
    for row in range(index_content_start, index_content_end):
        ws.row_dimensions[row].height = img_height*HEIGHT_FUDGE
    
    print("length of fn_list:", len(fn_list))
    for index, img_fn in enumerate(fn_list):
        if Path(img_fn).is_file():
            img = XSImage(img_fn)
            # print(f'image {index}')
            ws.add_image(img, img_col_letter+str(index+index_content_start))
        else:
            print(index, "not file")
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = al
    
    wb.save(f'{file.stem}-with-molecular-images.xlsx')
    

if __name__=="__main__":
    parser = argparse.ArgumentParser(description="Show SMILES")
    parser.add_argument("--f", "--file", help="input file")
    args = parser.parse_args()
    
    # args.file = "data/processed/test_100-pandas_tools_visulization-result.csv"
    # args.file = "data/processed/3CL-key-words-rcsb_pdb_custom_report_20221211-processed.csv"
    # args.file = "data/processed/Mpro-key-words-rcsb_pdb_custom_report_20221211-processed.csv"
    print(args)
    main(args)


















